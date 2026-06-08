#!/usr/bin/env python3
"""Собирает тестовый .xlsx «Сводный отчёт по ломбарду» из реальных майских
данных (та часть, что доступна), с реалистичной 2-уровневой шапкой и
заголовочными строками — чтобы проверить ридер read_svod_xlsx на структуре,
похожей на настоящую выгрузку 1С.
"""
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from parse_svod import parse_tables, records

# колонки, нужные расчёту (плоский ключ -> как разложить на 2 уровня шапки)
COLS = [
    "Территориальное образование",
    "Итого Вал/руб.",
    "Движение ТМЦ/Начальный остаток",
    "Движение ТМЦ/Конечный остаток",
    "Движение ТМЦ ВМ 585, гр./Начальный остаток",
    "Движение ТМЦ ВМ 585, гр./Конечный остаток",
    "Детализация/Выдано займов/Сумма, руб",
    "Детализация/Перемещено на торги/Сумма, руб",
    "Статистика/Количество заемщиков, чел./Конец периода",
]


def split_levels(key, nrows=2):
    parts = key.split("/")
    if len(parts) == 1:
        return [""] * (nrows - 1) + [parts[0]]
    # верхние уровни склеиваем в группу, последний — подпись
    return ["/".join(parts[:-1])] + [parts[-1]]


recs = records(*parse_tables(".work/svodbook.txt")[0])

wb = Workbook()
ws = wb.active
ws.title = "Май2026"

# заголовочные строки (как в выгрузке 1С)
ws.append(["Сводный отчет по ломбарду"])
ws.append(["за период Май 2026"])
ws.append(["Модули: Ломбард"])

# 2 строки шапки
hdr1, hdr2 = [], []
for key in COLS:
    a, b = split_levels(key, 2)
    hdr1.append(a)
    hdr2.append(b)
# имя филиала — в верхнюю строку шапки (она станет значением объединённой ячейки)
hdr1[0] = COLS[0]
hdr2[0] = ""
ws.append(hdr1)
ws.append(hdr2)

# данные
for r in recs:
    ws.append([r.get(k, "") for k in COLS])

# объединить вертикально ячейку имени в шапке (как обычно в выгрузке)
name_row1 = 4
ws.merge_cells(start_row=name_row1, start_column=1, end_row=name_row1 + 1, end_column=1)

out = ".work/sample_may.xlsx"
wb.save(out)
print(f"Сохранён тестовый файл: {out} ({len(recs)} филиалов, {len(COLS)} колонок)")
