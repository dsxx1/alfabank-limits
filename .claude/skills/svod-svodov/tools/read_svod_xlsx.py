#!/usr/bin/env python3
"""Чтение реальной выгрузки «Сводный отчёт по ломбарду» (.xlsx) из 1С.

Заголовок в выгрузке многоуровневый (группа / подгруппа / поле) с
объединёнными ячейками. Здесь он восстанавливается в плоские ключи вида
«Движение ТМЦ/Конечный остаток», «Детализация/Выдано займов/Сумма, руб» —
ровно те, что использует расчёт. Возвращает список словарей по подразделениям.
"""
from __future__ import annotations

import re
import sys

from openpyxl import load_workbook

NAME_COL = "Территориальное образование"
BRANCH_RE = re.compile(r"^\s*\d{3}-")


def _overlay(ws):
    """Карта (row,col)->значение с разворотом объединённых ячеек."""
    ov = {}
    for rng in ws.merged_cells.ranges:
        val = ws.cell(rng.min_row, rng.min_col).value
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                ov[(r, c)] = val
    return ov


def _cell(ws, ov, r, c):
    if (r, c) in ov:
        return ov[(r, c)]
    return ws.cell(r, c).value


def read_svod(path: str, sheet: str | None = None):
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    ov = _overlay(ws)

    # строка с подписью «Территориальное образование» — низ шапки
    label_row = None
    for r in range(1, ws.max_row + 1):
        if str(_cell(ws, ov, r, 1) or "").strip() == NAME_COL:
            label_row = r
            break
    if label_row is None:
        raise ValueError("Не найден заголовок «Территориальное образование»")

    # первая строка данных — первый филиал вида NNN-...
    data_start = None
    for r in range(label_row + 1, ws.max_row + 1):
        if BRANCH_RE.match(str(_cell(ws, ov, r, 1) or "")):
            data_start = r
            break
    if data_start is None:
        raise ValueError("Не найдены строки филиалов (NNN-...)")

    # шапка: все строки от label_row до data_start-1; ключ = join по уровням
    header_rows = list(range(label_row, data_start))
    ncol = ws.max_column
    keys = []
    for c in range(1, ncol + 1):
        parts = []
        for r in header_rows:
            v = _cell(ws, ov, r, c)
            v = str(v).strip() if v is not None else ""
            if v and (not parts or parts[-1] != v):
                parts.append(v)
        keys.append("/".join(parts))

    rows = []
    for r in range(data_start, ws.max_row + 1):
        name = _cell(ws, ov, r, 1)
        if not name or not BRANCH_RE.match(str(name)):
            continue
        rec = {}
        for c in range(1, ncol + 1):
            rec[keys[c - 1]] = _cell(ws, ov, r, c)
        rec[NAME_COL] = str(name).strip()
        rows.append(rec)
    return rows


if __name__ == "__main__":
    recs = read_svod(sys.argv[1])
    print(f"Прочитано филиалов: {len(recs)}")
    if recs:
        print("Ключи колонок:")
        for k in recs[0]:
            print("  -", k)
